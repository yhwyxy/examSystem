"""成绩导出。"""
from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from . import database
from . import question_loader



def format_student_answer_for_export(detail: dict[str, Any]) -> str:
    if detail.get("is_composite") and isinstance(detail.get("sub_results"), list):
        lines = []
        for s in detail["sub_results"]:
            sid = s.get("sub_question_id", "?")
            ans = s.get("student_answer", "")
            language = s.get("selected_language")
            prefix = f"[{sid}][{language}]" if language else f"[{sid}]"
            lines.append(f"{prefix} {ans}")
        return "\n".join(lines)
    ans = detail.get("student_answer", "")
    if isinstance(ans, list):
        return ",".join(str(x) for x in ans)
    if isinstance(ans, dict):
        return json.dumps(ans, ensure_ascii=False)
    return str(ans or "")


def format_score_note_for_export(detail: dict[str, Any]) -> str:
    if detail.get("is_composite") and isinstance(detail.get("sub_results"), list):
        parts = [
            f"{s.get('sub_question_id')}={s.get('final_score', s.get('score'))}/{s.get('max_score')}"
            for s in detail["sub_results"]
        ]
        return "复合题: " + "; ".join(parts)
    return str(detail.get("reason") or "")


def _answer_to_text(answer: Any) -> str:
    if isinstance(answer, dict):
        # 复合题答案 map：分行展示
        lines = [f"[{k}] {v}" for k, v in answer.items()]
        return "\n".join(lines)
    if isinstance(answer, list):
        return json.dumps(answer, ensure_ascii=False)
    if isinstance(answer, bool):
        return "正确" if answer else "错误"
    return "" if answer is None else str(answer)


def export_submissions_xlsx(paper_id: str | None = None) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ModuleNotFoundError as e:
        raise RuntimeError("导出 Excel 需要安装 openpyxl，请执行 pip install -r requirements.txt") from e

    rows = database.list_submissions(
        paper_id=paper_id, sort_by="submitted_at", order="desc", limit=10000, offset=0
    )

    question_ids: list[str] = []
    q_headers: list[str] = []
    if paper_id:
        try:
            questions = question_loader.get_question_list(paper_id)
            question_ids = [str(q["id"]) for q in questions]
            q_headers = [f"{q['id']}({q['score']}分)" for q in questions]
        except Exception:
            pass
    if not question_ids:
        seen: set[str] = set()
        for r in rows:
            detail_raw = r.get("grading_detail_json")
            if detail_raw:
                try:
                    details = json.loads(detail_raw) if isinstance(detail_raw, str) else detail_raw
                except Exception:
                    details = []
            else:
                details = []
            for d in details or []:
                qid = str(d.get("question_id") or "")
                if qid and qid not in seen:
                    seen.add(qid)
                    question_ids.append(qid)
                    q_headers.append(qid)

    wb = Workbook()
    ws = wb.active
    ws.title = "考试成绩"

    base_headers = [
        "ID", "姓名", "工号", "专业编码", "专业名称", "轮次", "部门",
        "客观题分", "主观题机器分", "主观题最终分", "总分", "复核状态", "提交时间",
    ]
    headers = base_headers + q_headers
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

    for r in rows:
        answers: dict[str, Any] = {}
        if r.get("answers_json"):
            try:
                answers = json.loads(r["answers_json"])
            except Exception:
                answers = {}
        if not answers and isinstance(r.get("answers"), dict):
            answers = r["answers"]

        details_by_qid: dict[str, Any] = {}
        detail_raw = r.get("grading_detail_json")
        if detail_raw:
            try:
                details = json.loads(detail_raw) if isinstance(detail_raw, str) else detail_raw
            except Exception:
                details = []
            for d in details or []:
                qid = str(d.get("question_id") or "")
                if qid:
                    details_by_qid[qid] = d

        if r.get("run_is_legacy"):
            round_label = "历史数据"
        elif r.get("round_no") is not None:
            round_label = f"第{r.get('round_no')}轮"
        else:
            round_label = ""

        base = [
            r.get("id"), r.get("name"), r.get("employee_id"),
            r.get("paper_id") or "", r.get("paper_name") or "",
            round_label,
            r.get("department"),
            r.get("objective_score"), r.get("subjective_score_machine"),
            r.get("subjective_score_final"), r.get("total_score"),
            r.get("review_status"), r.get("submitted_at"),
        ]
        ans_cols = []
        for qid in question_ids:
            d = details_by_qid.get(str(qid))
            if d and (d.get("is_composite") or d.get("sub_results")):
                cell = format_student_answer_for_export(d)
                note = format_score_note_for_export(d)
                if note:
                    cell = f"{cell}\n{note}"
                ans_cols.append(cell)
            else:
                ans_cols.append(_answer_to_text(answers.get(qid)))
        ws.append(base + ans_cols)

    for i, h in enumerate(headers, 1):
        letter = ws.cell(1, i).column_letter
        ws.column_dimensions[letter].width = min(max(len(str(h)) + 2, 10), 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
